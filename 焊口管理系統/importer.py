"""
importer.py — 把現有「焊口管制表」Excel 匯入資料庫(單一交易,適合大量資料)
支援 cp-129 / gl-05 / SE-37 / se-39 等不同欄位命名(以別名對應)。

CLI:
    python importer.py <xlsx路徑> <專案代號> "<專案名稱>" "<業主>"
"""
import sys
import datetime
import openpyxl

import db

JOINT_ALIASES = {
    "drawing_no": ["圖號", "DWGNO", "DWGNO.", "file_basename", "圖面編號"],
    "joint_no":   ["銲口編號", "焊口編號", "銲口號", "焊口號"],
    "serial":     ["流水號"],
    "size":       ["尺寸", "Size", "SIZE"],
    "thickness":  ["厚度", "THK"],
    "schedule":   ["SCH", "Schedule"],
    "material":   ["材質", "Material"],
    "weld_type":  ["銲接型式", "焊接型式", "銲接形式"],
    "db_factor":  ["係數"],
    "db_count":   ["DB數", "DB"],
    "shop_field": ["S/F", "預製S/現場F", "預製S/F", "預製/現場"],
    "weld_date":  ["配管完成日期", "組銲完成日期", "完成日期", "組焊完成日期"],
    "pipe_class": ["CLASS", "Class"],
    "category":   ["分類"],
    "subcontractor": ["承包商", "外包"],
    "nde_focus":  ["RT.PT檢驗焊口", "RT焊口", "RT.PT檢驗"],
    "nde_date":   ["RT.Date", "RT.PT檢驗日期", "RT日期", "RTDate"],
    "nde_percent": ["RT%", "RT％"],
    "nde_report": ["RT圖號", "報告號"],
    "remark":     ["備註", "NOTE", "REMARKS"],
}
DRAWING_ALIASES = {
    "drawing_no": ["file_basename", "圖號", "DWGNO", "DWGNO."],
    "serial":     ["流水號"],
    "system":     ["系統"],
    "line_no":    ["LINENO.", "LINENO", "LINENUMBER"],
    "pipe_class": ["Class", "CLASS"],
    "size":       ["Size", "SIZE", "尺寸"],
    "sheet_index": ["SheetIndex", "SH'TNO", "SHEETNO"],
    "num_sheets": ["NumSheets"],
    "rev":        ["版次"],
    "scan_date":  ["預製圖掃描日期", "預製掃描日期", "掃描日期"],
    "remark":     ["REMARKS", "備註"],
}


def normalize(s):
    if s is None:
        return ""
    return str(s).replace("\n", "").replace(" ", "").replace("　", "").strip()


def roc_to_iso(v):
    if v is None or v == "":
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    for sep in (".", "/", "-"):
        parts = s.split(sep)
        if len(parts) == 3:
            try:
                y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 1911:
                    y += 1911
                if 1950 <= y <= 2100 and 1 <= m <= 12 and 1 <= d <= 31:
                    return "%04d-%02d-%02d" % (y, m, d)
            except ValueError:
                pass
    return s


def build_map(header_row, aliases):
    norm = [normalize(h) for h in header_row]
    result = {}
    for field, names in aliases.items():
        for nm in names:
            key = normalize(nm)
            if key in norm:
                result[field] = norm.index(key)
                break
    return result


def cell(row, col_map, field):
    idx = col_map.get(field)
    if idx is None or idx >= len(row):
        return None
    v = row[idx]
    if isinstance(v, str):
        v = v.strip()
    return v if v != "" else None


def pick_sheet(wb, *predicates):
    for pred in predicates:
        for name in wb.sheetnames:
            if callable(pred):
                if pred(name):
                    return wb[name]
            elif name == pred:
                return wb[name]
    return None


def _s(v):
    return None if v is None else str(v).strip()


def _int(v):
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def import_weld_control(path, project_code, project_name, owner=None,
                        operator="importer", replace=True):
    db.init_db()
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    conn = db.connect()
    ex = conn.execute

    def scalar(sql, params=()):
        r = ex(sql, params).fetchone()
        return r[0] if r else None

    try:
        pid_row = ex("SELECT id FROM project WHERE code=?", (project_code,)).fetchone()
        if pid_row and replace:
            ex("DELETE FROM project WHERE id=?", (pid_row[0],))
            pid_row = None
        if pid_row:
            project_id = pid_row[0]
        else:
            project_id = ex("INSERT INTO project (code,name,owner) VALUES (?,?,?)",
                            (project_code, project_name, owner)).lastrowid

        for code, name, factor in db.DEFAULT_WELD_TYPES:
            ex("INSERT OR IGNORE INTO weld_type (project_id,code,name,factor) VALUES (?,?,?,?)",
               (project_id, code, name, factor))

        wt_sheet = pick_sheet(wb, "銲接型式", "銲接形式")
        if wt_sheet:
            for r in list(wt_sheet.iter_rows(values_only=True))[1:]:
                c = normalize(r[0]) if r and r[0] else None
                if c:
                    ex("INSERT OR IGNORE INTO weld_type (project_id,code,name,factor) VALUES (?,?,?,1)",
                       (project_id, c, str(r[1]) if len(r) > 1 and r[1] else c))
        wt_map = {row[1]: row[0] for row in
                  ex("SELECT id,code FROM weld_type WHERE project_id=?", (project_id,)).fetchall()}

        def weld_type_id(code):
            if not code:
                return None
            code = normalize(code)
            if code not in wt_map:
                wt_map[code] = ex("INSERT INTO weld_type (project_id,code,factor) VALUES (?,?,1)",
                                  (project_id, code)).lastrowid
            return wt_map[code]

        system_map, line_map, drawing_map, drawing_line = {}, {}, {}, {}

        def get_system(code):
            code = normalize(code)
            if not code:
                return None
            if code not in system_map:
                ex("INSERT OR IGNORE INTO system (project_id,code) VALUES (?,?)", (project_id, code))
                system_map[code] = scalar("SELECT id FROM system WHERE project_id=? AND code=?",
                                          (project_id, code))
            return system_map[code]

        def get_line(line_no, system_id):
            if not line_no:
                return None
            line_no = str(line_no).strip()
            if line_no not in line_map:
                ex("INSERT OR IGNORE INTO pipe_line (project_id,line_no,system_id) VALUES (?,?,?)",
                   (project_id, line_no, system_id))
                line_map[line_no] = scalar("SELECT id FROM pipe_line WHERE project_id=? AND line_no=?",
                                           (project_id, line_no))
            return line_map[line_no]

        dwg_sheet = pick_sheet(wb, "DWG NO.ALL",
                               lambda n: n.startswith("DWG NO.ALL") and "分析" not in n,
                               "DRAWING LIST",
                               lambda n: n.upper().startswith("DRAWING LIST") and "分析" not in n)
        n_dwg = 0
        if dwg_sheet:
            rows = list(dwg_sheet.iter_rows(values_only=True))
            cmap = build_map(rows[0], DRAWING_ALIASES) if rows else {}
            for r in rows[1:]:
                dno = cell(r, cmap, "drawing_no")
                if not dno:
                    continue
                dno = str(dno).strip()
                if dno in drawing_map:
                    continue
                sys_id = get_system(cell(r, cmap, "system"))
                line_id = get_line(cell(r, cmap, "line_no"), sys_id)
                drawing_map[dno] = ex(
                    "INSERT INTO drawing (project_id,drawing_no,serial_no,line_id,system_id,"
                    "pipe_class,size,sheet_index,num_sheets,current_rev,scan_date,remark) "
                    "VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                    (project_id, dno, _s(cell(r, cmap, "serial")), line_id, sys_id,
                     _s(cell(r, cmap, "pipe_class")), _s(cell(r, cmap, "size")),
                     _int(cell(r, cmap, "sheet_index")), _int(cell(r, cmap, "num_sheets")),
                     _s(cell(r, cmap, "rev")), roc_to_iso(cell(r, cmap, "scan_date")),
                     _s(cell(r, cmap, "remark")))).lastrowid
                drawing_line[dno] = line_id
                n_dwg += 1

        jt_sheet = pick_sheet(wb, "焊口編號明細",
                              lambda n: n.startswith("焊口編號明細") and "分析" not in n,
                              lambda n: n.startswith("焊口編號") and "分析" not in n)
        n_joint = n_insp = 0
        if jt_sheet:
            rows = list(jt_sheet.iter_rows(values_only=True))
            cmap = build_map(rows[0], JOINT_ALIASES) if rows else {}
            for r in rows[1:]:
                jno = cell(r, cmap, "joint_no")
                dno = cell(r, cmap, "drawing_no")
                if jno is None and dno is None:
                    continue
                jno = str(jno).strip() if jno is not None else ""
                drawing_id = None
                if dno is not None:
                    dno = str(dno).strip()
                    if dno not in drawing_map:
                        drawing_map[dno] = ex("INSERT OR IGNORE INTO drawing (project_id,drawing_no) VALUES (?,?)",
                                              (project_id, dno)).lastrowid
                    drawing_id = drawing_map[dno]
                wd = roc_to_iso(cell(r, cmap, "weld_date"))
                nde_focus = cell(r, cmap, "nde_focus")
                nde_date = roc_to_iso(cell(r, cmap, "nde_date"))
                report = _s(cell(r, cmap, "nde_report"))
                percent = _s(cell(r, cmap, "nde_percent"))
                try:
                    jid = ex(
                        "INSERT INTO weld_joint (project_id,drawing_id,line_id,joint_no,size,thickness,"
                        "schedule,material,weld_type_id,joint_category,db_factor,db_count,shop_field,"
                        "weld_date,subcontractor,nde_type,nde_percent,nde_date,nde_result,nde_report_no,"
                        "status,claim_status,remark) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
                        (project_id, drawing_id, drawing_line.get(dno), jno, _s(cell(r, cmap, "size")),
                         _s(cell(r, cmap, "thickness")), _s(cell(r, cmap, "schedule")),
                         _s(cell(r, cmap, "material")), weld_type_id(cell(r, cmap, "weld_type")),
                         _s(cell(r, cmap, "category")), _f(cell(r, cmap, "db_factor")) or 1,
                         _f(cell(r, cmap, "db_count")), _s(cell(r, cmap, "shop_field")), wd,
                         _s(cell(r, cmap, "subcontractor")),
                         "RT" if (nde_focus or nde_date) else None, percent, nde_date,
                         "合格" if nde_date else None, report,
                         "完成" if wd else "規劃", "未請款", _s(cell(r, cmap, "remark")))).lastrowid
                except Exception:
                    continue
                n_joint += 1
                if nde_date or nde_focus:
                    ex("INSERT INTO inspection (weld_joint_id,method,percent,inspect_date,result,report_no) "
                       "VALUES (?,?,?,?,?,?)",
                       (jid, "RT", percent, nde_date, "合格" if nde_date else None, report))
                    n_insp += 1

        bp_sheet = pick_sheet(wb, "請款期別")
        n_bp = 0
        if bp_sheet:
            seen = set()
            for r in list(bp_sheet.iter_rows(values_only=True))[1:]:
                if not r or len(r) < 2 or not r[1]:
                    continue
                code = str(r[1]).strip()
                if code and code not in seen:
                    seen.add(code)
                    ex("INSERT OR IGNORE INTO billing_period (project_id,code) VALUES (?,?)",
                       (project_id, code))
                    n_bp += 1

        ex("INSERT INTO audit_log (operator,action,entity,entity_id,summary) VALUES (?,?,?,?,?)",
           (operator, "IMPORT", "project", project_id,
            "匯入 %s:圖面 %d、焊口 %d、檢驗 %d、期別 %d" % (project_code, n_dwg, n_joint, n_insp, n_bp)))
        conn.commit()
    finally:
        conn.close()
        wb.close()

    return {"project_id": project_id, "drawings": n_dwg, "joints": n_joint,
            "inspections": n_insp, "billing_periods": n_bp,
            "systems": len(system_map), "lines": len(line_map)}


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("用法: python importer.py <xlsx> <專案代號> \"<名稱>\" [業主]")
        sys.exit(1)
    _path, _code, _name = sys.argv[1], sys.argv[2], sys.argv[3]
    _owner = sys.argv[4] if len(sys.argv) > 4 else None
    print("匯入完成:", import_weld_control(_path, _code, _name, _owner))
